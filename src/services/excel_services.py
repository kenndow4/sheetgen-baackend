from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

from src.dto.excel import ExcelRequestDto


def createExcelService(data: ExcelRequestDto):
    file_path = Path(data.filename)
    
   
    if file_path.exists():
        print(f"Archivo existente encontrado: {data.filename}")
        wb = load_workbook(data.filename)
        ws = wb.active
    else:
        print(f"Creando nuevo archivo: {data.filename}")
        wb = Workbook()
        ws = wb.active
        ws.append(data.columns)
        
      
        if data.header_style:
            apply_row_style(ws, 1, data.header_style)
    
    if data.column_configs:
        for col_config in data.column_configs:
            col_index = data.columns.index(col_config.name) + 1
            if col_config.width:
                ws.column_dimensions[ws.cell(1, col_index).column_letter].width = col_config.width
            
        
            if col_config.style:
                for row_num in range(2, ws.max_row + 1):
                    apply_cell_style(ws.cell(row_num, col_index), col_config.style)
    
    
    start_row = ws.max_row + 1
    for idx, row in enumerate(data.rows):
        ws.append(row)
        row_num = start_row + idx
        
        # Aplicar estilo a fila específica si existe
        if data.row_styles and str(idx) in data.row_styles:
            apply_row_style(ws, row_num, data.row_styles[str(idx)])
    
    wb.save(data.filename)
    print(f" Archivo guardado: {data.filename}")
    
    # 🔥 Devolver data completa para preview
    return {
        "filename": data.filename,
        "rows_added": len(data.rows),
        "columns": data.columns,
        "rows": data.rows,
        "header_style": data.header_style.dict() if data.header_style else None,
        "column_configs": [col.dict() for col in data.column_configs] if data.column_configs else None,
        "total_rows": len(data.rows)
    }


def apply_cell_style(cell, style):
    """Aplica estilo a una celda individual"""
    if style.bg_color:
        cell.fill = PatternFill(start_color=style.bg_color, end_color=style.bg_color, fill_type="solid")
    
    font_kwargs = {}
    if style.font_color:
        font_kwargs['color'] = style.font_color
    if style.bold:
        font_kwargs['bold'] = True
    if style.italic:
        font_kwargs['italic'] = True
    if style.font_size:
        font_kwargs['size'] = style.font_size
    
    if font_kwargs:
        cell.font = Font(**font_kwargs)
    
    if style.alignment:
        cell.alignment = Alignment(horizontal=style.alignment, vertical='center')


def apply_row_style(ws, row_num, style):
    """Aplica estilo a toda una fila"""
    for cell in ws[row_num]:
        apply_cell_style(cell, style)


